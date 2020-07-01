#! /usr/bin/python3
# -*- coding: utf-8 -*-   

"""

    使用本脚本之前，根据需要修改预设参数:
    SOURCE_CODE_PATH => 源码目录
    UNPACK_CODE_PATH => 解压目录
    XLSX_FILE        => 生成 xlsx 文件的地址
    LOG_FILE         => 生成 log 文件的地址
    CLOC_TIMEOUT     => 出现 timeout 的原因：
                        1. 执行 cloc 和 grep 的命令最长时间；
                        2. 输出过大阻塞管道
    BLACK_LIST       => 目前已知的较大的包，会卡死，需要跳过
    调整索引参数重新运行，目前对 xlsx 文件、log 文件会增量写入。

"""

import os
import time
import sys
import subprocess
import openpyxl

SOURCE_CODE_PATH = "/home/someone/CodeCount/Source"
UNPACK_CODE_PATH = "/home/someone/CodeCount/Unpack"
XLSX_FILE = "/home/someone/CodeCount/test.xlsx"
LOG_FILE = "/home/someone/CodeCount/test.log"
CLOC_TIMEOUT = 300
BLACK_LIST= [
    'thunderbird-52.2.1',
    'libreoffice-5.2.7',
    'firefox-esr-45.9.0esr',
    'gcc-5-5.4.1',
    'Unpack'
]

class TestCodeCount:
    """代码统计脚本

    说明：
        TestCodeCount 是专门为源码统计而编写的脚本, 使用本脚本之前，要在系统中安装 cloc 和 dpkg-dev 包。
        其中 cloc 统计代码数量， dpkg-dev 中 dpkg-source -x 用于解压源码包。

    日志：
        black_list=>源码文件过大，跳过该目录；
        list index out of range=>调用 cloc 没有生成 SUM: 结果，运行 grep 失败造成的，统计无效；
        subprocess returncode not 0=>返回值不为 0，可能执行命令失败

    参数:
        目录 cloc 统计中，timeout 设置为 300 秒，可根据需要调节。（例如 libboost 可以统计，但耗时很长，
        就超过120秒。）

    更新:
        20200628 first release by niyl
        20200629 split unpack and cloc function
        20200630 add from_index and to_index feature, or pass less than 0 to skip it
    """

    def __init__(self):
        """初始化"""
        self.cloc_cmd = "cloc --autoconf --by-file --exclude-dir .pc "
        self.failed_list = []
        self.black_list = BLACK_LIST

    def get_code_dirs(self, dir_path):
        """获取 dir_path 下所有目录

        :param dir_path：指定目录
        :return: 给定目录下的所有文件夹
        """
        all_files = os.listdir(dir_path)
        dirs = []
        # code-count 专门存放统计数据
        for file in all_files:
            if file in self.black_list:
                self.failed_list.append("black_list: {}".format(dir_path + '/' + file))
                continue
            if os.path.isdir(dir_path + '/' + file):
                dirs.append(os.path.join(dir_path, file))

        return dirs

    def test_get_code_dirs(self, dir_path):
        """get_code_dirs 的测试函数

        :param dir_path: 获取指定子目录的地址
        """
        dirs = self.get_code_dirs(dir_path)

        if dirs is None or len(dirs) == 0:
            print("directory is None")
            return

        print("test_get_code_dirs:")
        for file in dirs:
            print('dir list %s' % file)

    def get_dsc_files(self, dsc_files_path):
        """获取指定目录下的 dsc 文件

        :param dsc_files_path: dsc文件指定目录
        :return: dsc 文件列表
        """
        dsc_files = []
        for root, dirs, files in os.walk(dsc_files_path, topdown=False):
            for name in files:
                if name[-4:] == '.dsc':
                    dsc_files.append(os.path.join(root, name))
                else:
                    continue
        return dsc_files

    def test_get_dsc_files(self, dsc_files_path):
        """get_dsc_files的测试函数

        :param dsc_files_path: dsc文件指定目录
        """
        dsc_files = self.get_dsc_files(dsc_files_path)

        if dsc_files is None or len(dsc_files) == 0:
            print("dsc files is None")
            return

        print("find following dsc files:")
        for dsc in dsc_files:
            print(dsc_files_path + '/' + dsc)

    def build_workspace(self):
        """ 构建解压目录 """
        if os.path.isdir(UNPACK_CODE_PATH):
            return
        if not os.path.exists(UNPACK_CODE_PATH):
            os.mkdir(UNPACK_CODE_PATH)
        else:
            print("{} is not directory, exit".format(UNPACK_CODE_PATH))
            sys.exit(1)

    def subprocess_run_command(self, cmd, success_message, filename):
        """ 使用 subprocess 运行命令

        :param cmd: 运行的 shell 命令
        :param success_message: 成功时的输出信息
        :param filename: 是否要输出到文件
        """
        sub_proc = None
        try:
            if filename is None:
                sub_proc = subprocess.Popen(cmd, shell=True)
                sub_proc.wait(CLOC_TIMEOUT)
            else:
                sub_proc = subprocess.run(cmd,
                                          shell=True,
                                          check=True,
                                          stdout=subprocess.PIPE,
                                          stderr=subprocess.PIPE)
                with open(filename, 'ab+') as output_file:
                    output_file.write(sub_proc.stdout)
        except subprocess.TimeoutExpired as err:
            self.failed_list.append("subprocess command failed for: " + cmd)
            print(err)
        except subprocess.SubprocessError as err:
            self.failed_list.append("subprocess command failed for: " + cmd)
            print(err)
        except Exception as err:
            self.failed_list.append("subprocess command failed for: " + cmd)
            print(err)
        else:
            if not sub_proc and (sub_proc.returncode != '\0'):
                self.failed_list.append("subprocess returncode not 0: " + cmd)
            else:
                print(success_message)

    def unpack_source_packages(self, source_dir, unpack_dir, start_index, end_index):
        """解压源码包

        :param source_dir: 源码包存放的路径
        :param unpack_dir: 解压源码的路径
        :param start_index: 开始索引
        :param end_index: 结束索引
        :return:
        """

        dsc_files = self.get_dsc_files(source_dir)
        if dsc_files is None or len(dsc_files) == 0:
            print("No package need to unpack!")
            return

        if start_index > end_index:
            self.failed_list.append("unpack source package index error")

        unpack_index = 0
        dsc_index = 0

        dsc_files.sort()
        with open(LOG_FILE, 'a+') as log_file:
            log_file.write("dsc files list:\n")
            for dsc in dsc_files:
                dsc_index = dsc_index + 1
                print("dsc file: " + os.path.join(unpack_dir, dsc))
                log_file.write(os.path.join(unpack_dir, dsc) + '\n')
                if start_index > 0 and dsc_index < start_index:
                    continue
                if dsc_index > end_index > 0:
                    break
            log_file.write("===========================\n")

        for dsc in dsc_files:
            unpack_index = unpack_index + 1
            if start_index > 0 and unpack_index < start_index:
                continue
            if unpack_index > end_index > 0:
                break
            cmd = 'cd ' + unpack_dir + ';dpkg-source -x ' + dsc
            # print(cmd)
            self.subprocess_run_command(cmd, "unpack command success for " + dsc, None)

    def print_failed_list(self, failed_reason):
        """ 输出失败列表 """
        if not self.failed_list:
            return True

        with open(LOG_FILE, 'a+') as log_file:
            log_file.write(failed_reason)
            for item in self.failed_list:
                print("Error: " + item)
                log_file.write(item + '\n')
            log_file.write("===========================\n")
        return False

    def clear_none_directory(self, unpack_dir):
        """清理非目录文件

        由于异地解压，有些包可能会将源码文件拷贝，此处清理
        """
        all_files = os.listdir(unpack_dir)
        for file in all_files:
            path = os.path.join(unpack_dir, file)
            if not os.path.isdir(path):
                try:
                    os.remove(path)
                except:
                    print("Warning: clear none directory failed")

    def get_sum(self, subprocess_output):
        outputs = subprocess_output.decode('utf-8').strip().split()
        if outputs[0] == "SUM:":
            blank = int(outputs[1])
            commend = int(outputs[2])
            code = int(outputs[3])
            code_sum = blank + commend + code
            return code_sum
        return 0

    def generate_xlsx(self, unpack_codes, start_index, end_index):
        """ 生成 xlsx 文件 """
        codes_nums = len(unpack_codes)
        if codes_nums == 0:
            return
        if start_index > end_index:
            self.failed_list.append("generate xlsx index error")
            return
        xlsx_nodes = []
        index = 0
        unpack_codes.sort()
        for code_dir in unpack_codes:
            index = index + 1
            if start_index > 0 and index < start_index:
                continue
            if index > end_index > 0:
                break
            print("generate {}/{} node for {}".format(index, codes_nums, code_dir))
            package_name = "NoName"
            cloc_sum = 0
            package_maintainer = "NoMaintainer"
            package_homepage = "NoHomePage"
            package_description = "NoDescription"

            try:
                cmd1 = self.cloc_cmd + code_dir
                cmd2 = "grep SUM:"
                cloc_proc = subprocess.Popen(cmd1, stdout=subprocess.PIPE, shell=True)
                cloc_proc2 = subprocess.Popen(cmd2, stdin=cloc_proc.stdout, stdout=subprocess.PIPE, shell=True)
                cloc_proc2.wait(CLOC_TIMEOUT)
                cloc_sum_output = cloc_proc2.stdout.readlines()[0]
                cloc_sum = self.get_sum(cloc_sum_output)
            except Exception as e:
                print("{} for {}\n".format(e, code_dir))
                self.failed_list.append("{} for {}".format(e, code_dir))
                cloc_sum = 0
            control_file = code_dir + '/debian/control'
            if not os.path.exists(control_file):
                self.failed_list.append("get control file failed: " + control_file)
                continue
            with open(control_file, 'r+') as file:
                for line in file.readlines():
                    line = line.strip('\n')
                    if line.startswith("Source:"):
                        package_name = line.split("Source:")[1].strip()

                    if line.startswith("Maintainer:"):
                        package_maintainer = line.split("Maintainer:")[1].strip()

                    if line.startswith("Homepage:"):
                        package_homepage = line.split("Homepage:")[1].strip()

                    if line.startswith("Description:"):
                        package_description = line.split("Description:")[1].strip()

                    if line.startswith("Description:"):
                        break
            xlsx_nodes.append([package_name,
                               cloc_sum,
                               package_maintainer,
                               package_homepage,
                               package_description])

        if not os.path.exists(XLSX_FILE):
            data = openpyxl.Workbook()
            sheetnames = data.get_sheet_names()
            sheet = data.get_sheet_by_name(sheetnames[0])
            sheet['A1'] = "Package Name"
            sheet['B1'] = "Cloc"
            sheet['C1'] = "Package Maintainer"
            sheet['D1'] = "Package Homepace"
            sheet['E1'] = "Package Description"
        else:
            # 追加写入
            data = openpyxl.load_workbook(XLSX_FILE)
            sheet = data.active
        for node in xlsx_nodes:
            sheet.append(node)
        data.save(filename=XLSX_FILE)

    def start_cloc(self, option="NoOption", start_index=-1, end_index=-1):
        print("%s start ..." % option)
        self.failed_list = []
        if option == "unpack":
            # unpack
            self.build_workspace()
            self.unpack_source_packages(SOURCE_CODE_PATH, UNPACK_CODE_PATH, start_index, end_index)
            self.clear_none_directory(UNPACK_CODE_PATH)
            res = self.print_failed_list("unpack failed list:\n")
            if not res:
                return
        elif option == "calcu":
            # calculate
            unpack_codes = self.get_code_dirs(UNPACK_CODE_PATH)
            self.generate_xlsx(unpack_codes, start_index, end_index)
            res = self.print_failed_list("generate xlsx failed list:\n")
            if not res:
                return
        else:
            print("option argument error")
        # debug
        print("%s end ..." % option)

    def print_help(self):
        print("Usage:")
        print("code-count.py option [from_index to_index]")
        print("option can be unpack or calcu")

if __name__ == '__main__':
    # 参数模式
    cc = TestCodeCount()
    argc = len(sys.argv)

    if argc != 2 and argc != 4:
        cc.print_help()
        sys.exit(1)
    option_type = sys.argv[1]
    from_index = -1
    to_index = -1
    if argc == 4:
        try:
            from_index = int(sys.argv[2])
            to_index = int(sys.argv[3])
        except:
            print("Warning: index argument error, use -1 as default")
            from_index = -1
            to_index = -1
    cc.start_cloc(option_type, from_index, to_index)

    # =======================================

    # 手动模式
    # cc.start_cloc("unpack", 1, 10)
    # time.sleep(5)
    # cc.start_cloc("calcu", 1, 10)
    # time.sleep(5)
    # cc.start_cloc("calcu", 16, 30)
    # time.sleep(5)
    # cc.start_cloc("calcu", 31, 45)
    # time.sleep(5)
    # cc.start_cloc("calcu", 46, 60)
    # time.sleep(5)
    # cc.start_cloc("calcu", 801, 1000)
    # time.sleep(5)
    # cc.start_cloc("calcu", 1001, 1200)
